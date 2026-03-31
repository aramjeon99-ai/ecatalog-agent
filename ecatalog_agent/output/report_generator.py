from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ecatalog_agent.models.state import AgentState


def _write_row(ws, row_idx: int, values: list[Any]) -> None:
    for col_idx, v in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=v)


def generate_review_report(
    state: AgentState,
    *,
    output_dir: str | Path,
) -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    request_id = state.record.request_id
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"review_report_{request_id}_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Review"

    ws["A1"] = "e-Catalog Agent Review Report"

    _write_row(
        ws,
        3,
        [
            "Request ID",
            state.record.request_id,
            "Model",
            state.record.model_name,
            "Maker",
            state.record.maker_name,
        ],
    )

    outcome = state.final_decision.outcome if state.final_decision else None
    _write_row(
        ws,
        4,
        ["Outcome", outcome, "Decided At", state.final_decision.decided_at if state.final_decision else ""],
    )

    step_results = [
        state.step0_result,
        state.step1_result,
        state.step2_result,
        state.step3_result,
        state.step4_result,
        state.step5_result,
    ]

    start_row = 6
    _write_row(ws, start_row, ["Step", "Status", "Confidence", "Flags", "Key Details"])

    r = start_row + 1
    for sr in step_results:
        if sr is None:
            continue
        flags = ", ".join([f.code for f in sr.flags_raised]) if sr.flags_raised else ""
        details = sr.details or {}
        # Compact "key details" for readability.
        key_details = ", ".join([f"{k}={v}" for k, v in list(details.items())[:4]])
        _write_row(ws, r, [sr.step_name, sr.status, float(sr.confidence), flags, key_details])
        r += 1

    # Autosize columns roughly.
    for col_idx in range(1, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28

    if state.final_decision and state.final_decision.rejection_summary:
        ws["A" + str(r + 2)] = "Rejection Summary"
        ws["B" + str(r + 2)] = state.final_decision.rejection_summary

    wb.save(str(out_path))
    return out_path

